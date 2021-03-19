# Still needs work but so far runs good
# this is the one

# created on 16-3-2021

import PyPDF2  # to handle PDFs
import imghdr  # to handle Images
import os
import smtplib
import time

import openpyxl
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText 
print('emailer code program\nThat works')
wb = openpyxl.load_workbook('D://raima//Desktop//Emailer//emails 102.xlsx')
ws = wb['Sheet1']
toSend = []
toSend_names = []
for i in range(2, ws.max_row + 1):
    email_receiver = ws.cell(row=i, column=1).value
    toSend.append(email_receiver)
    name = ws.cell(row=i, column=2).value
    toSend_names.append(name)
print('___________________________________________')
print('total emails = ' + str(len(toSend)))
# username = str(input('Your Username:'))
# password = str(input('Your Password:'))
subject = 'Are you interested in Veterinary Instruments & Farrier Tools'


# Create message container - the correct MIME type is multipart/alternative.
# msg = MIMEMultipart()
print('starting from ' + str(toSend[0]))


def send_mail():
    for m in range(len(toSend)):
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        # server.starttls()
        # server.login(username, password)-------
        server.login('snb.khurram@gmail.com', 'khurram.saeed03360736736')
        email = EmailMessage()
        email['From'] = 'snb.khurram@gmail.com'
        email['To'] = toSend[m]
        email['Subject'] = subject
        text = '''Dear ''' + str(toSend_names[m]) + '''\n
            I would like to introduce our company SAEED NAVEED BROTHERS \n
            SAEED NAVEED BROTHERS is a Veterinary Instruments, Farrier Tools, Equestrian Products manufacturing and export company in Pakistan.\n
            (Our catalog is attached; we can also send you our printed catalog)\n
            A list of product categories includes:\n
            •   Bits, Spurs, Stirrups\n
            •	Equestrian Health Products\n
            •	Veterinary Instruments\n
            •	Farrier Tools\n
            •	Hoof knives, Folding knives, Custom knives\n
            •	Tongs\n
            •	Clinchers, Adjustable Clinchers\n
            •	Hackamores\n
            •	Eggbutt Bits\n
            •	Cheek Mouth Pieces\n  
            •	Castration Forceps\n
            •	Hoof Testers\n
            •	Pet Grooming Shears\n
            •	Equestrian Leather Products\n
            •	Riding Gloves & Chaps\n
             \n
            If you need any sample to check our quality. We will be happy to send it to you.\n
            We do Business World Wide, looking forward to doing good business with you.\n
            \n
            Best regards\n
            \n
             KHURRAM SAEED\n
            SALES DIRECTOR\n
             \n
            SAEED NAVEED BROTHERS LTD\n
            MANUFACTURER & EXPORTER OF VETERINARY INSTRUMENTS, EQUESTRIAN HARDWARE AND FARRIER TOOLS\n
             \n
            Gmail:   snbrothers@gmail.com\n
            Email:   info@snbrothers.com\n   
            Cell:   +92 321 611 6509\n
            Phone:   +92 52 355 4004\n
            YouTube: https://www.youtube.com/results?search_query=saeed+naveed+brothers\n
            Instagram: https://www.instagram.com/snb.company.1971/\n
            Facebook:   www.facebook.com/saeednaveedbrothers\n
            Alibaba:   www.snbrothers.trustpass.alibaba.com\n
            PayPal:   www.paypal.me/SaeedAkhtar\n
            Address:   P.O. Box 1322, 11c Nishter Road, SIE, Sialkot-51310 (Pakistan)\n
            '''
        # html = '''<p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",
        # sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;'><span
        # style='font-family:"Calibri",sans-serif;color:black;'>Dear ''' + str(toSend_names[m]) + '''</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;text-align:start;'><span style='font-family:"Calibri",sans-serif;color:black;'>I would like to introduce our company&nbsp;</span><strong><span style='font-size:19px;font-family:"Calibri",sans-serif;color:black;'>SAEED NAVEED BROTHERS&nbsp;</span></strong></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;text-align:start;'><span style='font-family:"Calibri Light",sans-serif;color:black;'>SAEED NAVEED BROTHERS is a Veterinary Instruments, Farrier Tools, Equestrian Products manufacturing and export company in Pakistan.</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;text-align:start;'><span style='font-size:15px;font-family:"Calibri Light",sans-serif;color:black;'>&nbsp;</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;text-align:start;'><span style='font-family:"Calibri Light",sans-serif;color:red;'>(Our catalog is attached; we can also send you our printed catalog)</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;text-align:start;'><span style='font-size:15px;font-family:"Calibri Light",sans-serif;color:black;'>&nbsp;</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;text-align:start;'><span style='font-family:"Calibri Light",sans-serif;color:black;'>A list of product categories includes:</span></p>
        # <ul style="margin-bottom:0in;text-align:start;" type="disc">
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Bits, Spurs, Stirrups</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Equestrian Health Products</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Veterinary Instruments</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Farrier Tools</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Hoof knives, Folding knives, Custom knives</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Tongs</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Clinchers, Adjustable Clinchers</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Hackamores</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Eggbutt Bits</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Cheek Mouth Pieces &nbsp;</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Castration Forceps</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Hoof Testers</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Pet Grooming Shears</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Equestrian Leather Products</span></li>
        #     <li style='margin-top:0in;margin-right:0in;margin-bottom:8.0pt;margin-left:0in;line-height:normal;font-size:15px;font-family:"Calibri",sans-serif;color:black;'><span style='font-family:"Calibri Light",sans-serif;'>Riding Gloves &amp; Chaps</span></li>
        # </ul>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;text-align:start;'><span style='font-family:"Calibri Light",sans-serif;color:black;'>&nbsp;</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;text-align:start;'><span style='font-family:"Calibri Light",sans-serif;color:black;'>If you need any sample to check our quality. We will be happy to send it to you.</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;text-align:start;'><span style='font-size:15px;font-family:"Calibri Light",sans-serif;color:black;'>We do Business World Wide,&nbsp;</span><span style='font-family:"Calibri Light",sans-serif;color:black;'>looking forward to doing good business with you.</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;text-align:start;'><span style='font-size:15px;font-family:"Calibri Light",sans-serif;color:black;'>&nbsp;</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;text-align:start;'><span style='font-family:"Calibri Light",sans-serif;color:black;'>Best regards</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;text-align:start;'><span style='font-size:15px;font-family:"Garamond",serif;color:#500050;'><br>&nbsp; KHURRAM SAEED</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;background:white;text-align:start;'><span style='font-size:13px;font-family:"Garamond",serif;color:#500050;'>SALES DIRECTOR</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;background:white;text-align:start;'><span style='font-size:13px;font-family:"Arial",sans-serif;color:#500050;'>&nbsp;</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;background:white;text-align:start;'><span style='font-size:27px;font-family:"Garamond",serif;color:#444444;'>SAEED NAVEED BROTHERS&nbsp;</span><span style='font-size:15px;font-family:"Garamond",serif;color:#444444;'>LTD</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;background:white;text-align:start;'><span style='font-size:13px;font-family:"Garamond",serif;color:#666666;'>MANUFACTURER &amp; EXPORTER OF VETERINARY INSTRUMENTS,&nbsp;EQUESTRIAN HARDWARE AND FARRIER TOOLS</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;background:white;text-align:start;'><span style='font-size:13px;font-family:"Arial",sans-serif;color:#500050;'>&nbsp;</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;background:white;text-align:start;'><span style='font-size:15px;font-family:"Courier New";color:#666666;'>Gmail: &nbsp;&nbsp;</span><u><span style='font-size:15px;font-family:"Courier New";color:#1155CC;'><a href="mailto:snbrothers@gmail.com" target="_blank"><span style="color:#1155CC;">snbrothers@gmail.com</span></a></span></u></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;background:white;text-align:start;'><span style='font-size:15px;font-family:"Courier New";color:#666666;'>Email: &nbsp;&nbsp;</span><u><span style='font-size:15px;font-family:"Courier New";color:#1155CC;'><a href="mailto:info@snbrothers.com" target="_blank"><span style="color:#1155CC;">info@snbrothers.com</span></a></span></u><span style='font-size:15px;font-family:"Courier New";color:#666666;'>&nbsp; &nbsp;<br>Cell: &nbsp; +92 321 611 6509</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;background:white;text-align:start;'><span style='font-size:15px;font-family:"Courier New";color:#666666;'>Phone: &nbsp; +92 52 355 4004</span></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;background:white;text-align:start;'><span style='font-size:15px;font-family:"Courier New";color:#666666;'>YouTube:&nbsp;</span><u><span style='font-size:15px;font-family:"Courier New";color:blue;'><a href="https://www.youtube.com/results?search_query=saeed+naveed+brothers">https://www.youtube.com/results?search_query=saeed+naveed+brothers</a></span></u></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;background:white;text-align:start;'><span style='font-size:15px;font-family:"Courier New";color:#767171;'>Instagram:&nbsp;</span><u><span style='font-size:15px;font-family:"Courier New";color:blue;'><a href="https://www.instagram.com/snb.company.1971/">https://www.instagram.com/snb.company.1971/</a></span></u><u><span style='font-size:21px;font-family:"Courier New";color:#666666;'><br></span></u><span style='font-size:15px;font-family:"Courier New";color:#666666;'>Facebook: &nbsp;&nbsp;</span><u><span style='font-size:15px;font-family:"Courier New";color:#1155CC;'><a href="http://www.facebook.com/saeednaveedbrothers" target="_blank"><span style="color:#1155CC;">www.facebook.com/saeednaveedbrothers</span></a></span></u><span style='font-size:15px;font-family:"Courier New";color:#666666;'><br>Alibaba: &nbsp;&nbsp;</span><u><span style='font-size:15px;font-family:"Courier New";color:#1155CC;'><a href="http://www.snbrothers.trustpass.alibaba.com/" target="_blank"><span style="color:#1155CC;">www.snbrothers.trustpass.alibaba.com</span></a></span></u><span style='font-size:15px;font-family:"Courier New";color:#666666;'><br>PayPal: &nbsp;&nbsp;</span><u><span style='font-size:15px;font-family:"Courier New";color:#1155CC;'><a href="http://www.paypal.me/SaeedAkhtar" target="_blank"><span style="color:#1155CC;">www.paypal.me/SaeedAkhtar</span></a></span></u></p>
        # <p style='margin-right:0in;margin-left:0in;font-size:15px;font-family:"Calibri",sans-serif;margin-top:0in;margin-bottom:.0001pt;line-height:107%;margin:0in;background:white;text-align:start;'><span style='font-size:15px;font-family:"Courier New";color:#666666;'>Address: &nbsp; P.O. Box 1322, 11c Nishter Road, SIE, Sialkot-51310 (Pakistan)</span></p>'''

        # part1=MIMEText(text,'plain')
        # part2=MIMEText(html,'html')
        # msg.attach(part1)
        # msg.attach(part2)
        email.set_content(text)
        # Attaching pdf

        with open('D://raima//Desktop//Business//snb-main.pdf', 'rb') as f:
            file_data = f.read()
            file_name = 'SNB Main Catalog'
        email.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)

        with open('D://raima//Desktop//Business//snb-vet.pdf', 'rb') as f:
            file_data = f.read()
            file_name = 'SNB Veterinary Catalog'
        email.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)

        with open('D://raima//Desktop//Business//banner.jpg', 'rb') as image:
            image_data = image.read()
            image_type = imghdr.what(image.name)
            image_name = 'SNB Banner'
        email.add_attachment(image_data, maintype='image', subtype=image_type, filename=image_name)

        server.send_message(email)
        print(str(m + 1) + '. sent to ' + str(toSend[m]))
        time.sleep(10)


print('login...')
send_mail()
print('All ' + str(len(toSend)) + ' are sent Successfully')
